#!/usr/bin/env python3
import sys, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_excel(json_file, output_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        employees = json.load(f)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'รายงานพนักงาน'

    headers = [
        'รหัสพนักงาน','คำนำหน้า','ชื่อจริง','นามสกุล','ชื่อจริง (EN)','นามสกุล (EN)',
        'เลขบัตรประชาชน','แผนก','ตำแหน่ง','ประเภทพนักงาน','อีเมล','เบอร์โทร',
        'ผู้ติดต่อฉุกเฉิน','เบอร์ติดต่อฉุกเฉิน','ที่อยู่','แขวง/ตำบล','เขต/อำเภอ',
        'จังหวัด','รหัสไปรษณีย์','วันที่เริ่มทำงาน YYYYMMDD (ค.ศ.)','เงินเดือน',
        'บัญชีเงินเดือนที่บันทึก','ประกันสังคม','ยอดหัก ณ ที่จ่าย (ภ.ง.ด.1)',
        'ช่องทางรับเงิน','เลขที่บัญชี','สถานะ','ทีม','username','password'
    ]

    hf = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = hf
        cell.font = hfont
        cell.alignment = center

    status_map = {'active':'พนักงาน','inactive':'ลาออก','suspended':'ระงับ'}

    for e in employees:
        start = e.get('start_date','')
        if start:
            try:
                p = start.split('-')
                start = f"{p[2]}/{p[1]}/{p[0]}"
            except: pass

        row = [
            e.get('employee_id',''),
            e.get('title',''),
            e.get('first_name',''),
            e.get('last_name',''),
            e.get('first_name_en','') or '',
            e.get('last_name_en','') or '',
            e.get('id_card_number','') or '',
            e.get('dept_name','') or '',
            e.get('position','') or '',
            e.get('employee_type','รายเดือน') or 'รายเดือน',
            e.get('email','') or '',
            e.get('phone','') or '',
            e.get('emergency_contact','') or '',
            e.get('emergency_phone','') or '',
            e.get('address','') or '',
            e.get('sub_district','') or '',
            e.get('district','') or '',
            e.get('province','') or '',
            e.get('postal_code','') or '',
            start,
            float(e.get('base_salary',0) or 0),
            e.get('salary_account','') or '',
            e.get('social_security_status','ขึ้นทะเบียนประกันสังคม') or '',
            float(e.get('withholding_tax',0) or 0),
            e.get('payment_channel','') or '',
            e.get('bank_account','') or '',
            status_map.get(e.get('status','active'),'พนักงาน'),
            e.get('team_name','') or '',
            e.get('username',''),
            '',
        ]
        ws1.append(row)

    col_widths = [12,10,15,15,15,15,18,14,14,12,28,14,18,16,30,14,14,12,10,22,12,28,30,15,25,16,10,18,14,10]
    for i, cw in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = cw
    ws1.freeze_panes = 'A2'
    ws1.row_dimensions[1].height = 35

    # Sheet 2: เวลาเข้างาน-ออกงาน
    ws2 = wb.create_sheet('เวลาเข้างาน-ออกงาน')
    ws2.append([None,None,None,None,None])
    ws2.append([None,None,None,None,None])
    ws2.append(['ชื่อตำแหน่ง (position_name)','เวลาเข้างาน (time_in)','เวลาออกงาน (time_out)','สายได้ไม่เกิน/นาที (late_allowance_mins)','วันทำงาน (work_days)'])
    ws2.append(['ผู้จัดการ','08.30','17.30',15,'จันทร์-ศุกร์'])
    ws2.append(['ผู้ปฏิบัติการ','08.30','17.30',15,'จันทร์-ศุกร์'])
    ws2.append(['ผู้บริหาร (Owner)','08.30','17.30',0,'จันทร์-ศุกร์'])

    # Sheet 3: ประเภทการลา
    ws3 = wb.create_sheet('ประเภทการลา')
    ws3.append([None]*8)
    ws3.append([None]*8)
    ws3.append(['รหัสการลา (leave_code)','ชื่อประเภทการลา (leave_name)','จำนวนวัน/ปี (default_days)','หักเงินเดือน (is_deduct_salary)','ต้องแนบเอกสาร (require_doc_days)','สามารถลาเป็นชั่วโมงได้ไหม','ทบปีถัดไป (is_carry_forward)','หมายเหตุ / เงื่อนไข (description)'])
    for lt in [('SL','ลาป่วย',30,False,3,False,False,'ตามกฎหมายแรงงาน'),('PL','ลากิจ',3,False,0,True,False,'สำหรับธุระจำเป็น'),('AL','ลาพักร้อน',6,False,0,False,True,'สำหรับพนักงานที่ผ่านโปร'),('ML','ลาคลอด',98,False,1,False,False,'ได้รับค่าจ้าง 45 วันแรก'),('UL','ลาไม่รับค่าจ้าง',0,True,0,False,False,'หักเงินเดือน')]:
        ws3.append(list(lt))

    # Sheet 4: ทีม
    ws4 = wb.create_sheet('ทีม')
    ws4.append(['รหัสทีม (team_code)','ชื่อทีม/สาขา (team_name)','รหัสพนักงานของหัวหน้าทีม(ถ้ามี) (manager_employee_id)','พื้นที่/จังหวัด (location)','เป้ายอดขายต่อเดือน (monthly_sales_target)','สถานะ (is_active)'])

    wb.save(output_file)
    print("OK")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        sys.exit(1)
    export_excel(sys.argv[1], sys.argv[2])
