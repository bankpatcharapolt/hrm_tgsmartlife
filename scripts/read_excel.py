#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, json, os

def read_excel(filepath):
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({'error': 'openpyxl not installed. Run: pip install openpyxl'}, ensure_ascii=False))
        return

    if not os.path.exists(filepath):
        print(json.dumps({'error': 'File not found: ' + filepath}, ensure_ascii=False))
        return

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(json.dumps({'error': 'Cannot open file: ' + str(e)}, ensure_ascii=False))
        return

    # หา sheet รายงานพนักงาน — ลอง exact match ก่อน ถ้าไม่เจอใช้ sheet แรก
    ws = None
    target_names = ['รายงานพนักงาน', 'employee', 'employees', 'staff', 'พนักงาน']
    
    for name in wb.sheetnames:
        if name.strip() == 'รายงานพนักงาน':
            ws = wb[name]
            break
    
    if ws is None:
        # ลอง fuzzy match
        for name in wb.sheetnames:
            n_lower = name.lower().strip()
            if any(t in n_lower for t in ['พนักงาน', 'employee', 'staff']):
                ws = wb[name]
                break
    
    if ws is None:
        # fallback: ใช้ sheet แรก
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.values)
    if not rows or len(rows) < 2:
        print(json.dumps([], ensure_ascii=False))
        return

    # skip header (row 1)
    data_rows = rows[1:]
    result = []
    
    for row in data_rows:
        if not any(v is not None for v in row):
            continue  # ข้ามแถวว่างทั้งหมด
        
        clean = []
        for cell in row:
            if cell is None:
                clean.append(None)
            elif hasattr(cell, 'strftime'):
                # datetime object
                clean.append(cell.strftime('%d/%m/%Y'))
            elif isinstance(cell, bool):
                clean.append(cell)
            elif isinstance(cell, (int, float)):
                clean.append(cell)
            elif isinstance(cell, str):
                clean.append(cell.strip())
            else:
                clean.append(str(cell).strip())
        result.append(clean)

    # encode ให้ stdout รองรับ UTF-8
    output = json.dumps(result, ensure_ascii=False, default=str)
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout.buffer.write(output.encode('utf-8'))
        sys.stdout.buffer.write(b'\n')
    else:
        print(output)

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Usage: read_excel.py <file.xlsx>'}, ensure_ascii=False))
        sys.exit(1)
    read_excel(sys.argv[1])
